from __future__ import annotations

import json
import math
import re
import shutil
import threading
import uuid
from datetime import datetime
from pathlib import Path
from typing import NamedTuple

from openpyxl import load_workbook
from PIL import Image, ImageChops, ImageFilter
from psd_tools import PSDImage

from ..config import OUTPUTS_DIR
from ..db import dump_json, get_conn
from .asset_service import cutout_asset
from .excel_service import _parse_group_columns, _extract_wps_embedded_images
from .psd_service import _safe_bbox, build_layer_map, composite_without_layers


# ---------------------------------------------------------------------------
# Image placement helpers
# ---------------------------------------------------------------------------

class SlotResult(NamedTuple):
    img: Image.Image
    paste_x: int
    paste_y: int


def _trim_transparent(img: Image.Image) -> Image.Image:
    """
    Crop away fully-transparent border pixels so the bounding box of the
    actual content is used for scaling rather than the full canvas size
    (e.g. an 800x800 rembg output where the subject occupies only 200x300
    would otherwise be shrunk to fit 800x800 into a small slot).
    Falls back to the original image if the result would be empty.
    """
    if img.mode != "RGBA":
        return img
    bbox = img.split()[3].getbbox()  # alpha channel bounding box
    if bbox is None:
        return img  # fully transparent – keep as-is
    cropped = img.crop(bbox)
    if cropped.width == 0 or cropped.height == 0:
        return img
    return cropped


def _fit_image_to_slot(img: Image.Image, slot_x: int, slot_y: int, slot_w: int, slot_h: int) -> SlotResult:
    """
    Scale img (preserving aspect ratio) to fit inside slot_w x slot_h,
    then center it at (slot_x, slot_y).  Both enlarging and shrinking allowed.
    Transparent borders are trimmed before scaling so the visible subject
    fills the slot rather than the whole canvas.
    """
    if slot_w <= 0 or slot_h <= 0:
        return SlotResult(img, slot_x, slot_y)

    img = _trim_transparent(img)
    img_w, img_h = img.size
    if img_w == 0 or img_h == 0:
        return SlotResult(img, slot_x, slot_y)

    scale = min(slot_w / img_w, slot_h / img_h)
    new_w = max(1, int(img_w * scale))
    new_h = max(1, int(img_h * scale))

    resized = img.resize((new_w, new_h), Image.LANCZOS)

    paste_x = slot_x + (slot_w - new_w) // 2
    paste_y = slot_y + (slot_h - new_h) // 2
    return SlotResult(resized, paste_x, paste_y)


def _grid_layout(
    images: list[Image.Image],
    bbox_x: int,
    bbox_y: int,
    bbox_w: int,
    bbox_h: int,
) -> list[SlotResult]:
    """
    Lay out N images in an auto grid inside the given bounding box.
    cols = ceil(sqrt(N)), rows = ceil(N/cols).
    Each cell uses _fit_image_to_slot.
    """
    n = len(images)
    if n == 0:
        return []

    cols = math.ceil(math.sqrt(n))
    rows = math.ceil(n / cols)
    cell_w = bbox_w // cols
    cell_h = bbox_h // rows

    results: list[SlotResult] = []
    for idx, img in enumerate(images):
        col = idx % cols
        row = idx // cols
        cell_x = bbox_x + col * cell_w
        cell_y = bbox_y + row * cell_h
        results.append(_fit_image_to_slot(img, cell_x, cell_y, cell_w, cell_h))

    return results


def _flatten_leaf_layers(group, result: list) -> None:
    """Collect only non-group PSD layers in traversal order."""
    for layer in group:
        if layer.is_group():
            _flatten_leaf_layers(layer, result)
            continue
        result.append(layer)


def _render_layer_aware_canvas(
    *,
    ordered_layers: list,
    layer_id_by_object_id: dict[int, str],
    replace_layer_ids: set[str],
    replacement_slots: dict[str, SlotResult],
    canvas_size: tuple[int, int],
) -> Image.Image:
    """
    Render canvas by replaying PSD leaf layers in original stack order.
    Replace-layer slots are injected at each target layer position.
    """
    canvas = Image.new("RGBA", canvas_size, (0, 0, 0, 0))

    for layer in ordered_layers:
        if not getattr(layer, "visible", True):
            continue

        layer_id = layer_id_by_object_id.get(id(layer))
        if layer_id in replace_layer_ids:
            slot = replacement_slots.get(layer_id)
            if slot is not None:
                canvas.alpha_composite(slot.img, (slot.paste_x, slot.paste_y))
            continue

        try:
            rendered = layer.composite(force=True)
        except TypeError:
            rendered = layer.composite()

        if not isinstance(rendered, Image.Image):
            continue

        rendered_rgba = rendered.convert("RGBA")
        if rendered_rgba.size == canvas.size:
            canvas.alpha_composite(rendered_rgba, (0, 0))
            continue

        x1, y1, _, _ = _safe_bbox(layer)
        canvas.alpha_composite(rendered_rgba, (x1, y1))

    return canvas


def _build_slots_mask(
    canvas_size: tuple[int, int],
    slots: list[SlotResult],
    *,
    dilate_px: int = 0,
) -> Image.Image:
    """
    Build editable mask from real replacement alpha rather than full slot rectangles.
    A tiny dilation absorbs interpolation/antialias edge drift.
    """
    width, height = canvas_size
    mask = Image.new("L", (width, height), 0)
    for slot in slots:
        slot_w, slot_h = slot.img.size
        if slot_w <= 0 or slot_h <= 0:
            continue
        alpha = slot.img.split()[-1]
        x1 = max(0, slot.paste_x)
        y1 = max(0, slot.paste_y)
        x2 = min(width, slot.paste_x + slot_w)
        y2 = min(height, slot.paste_y + slot_h)
        if x2 <= x1 or y2 <= y1:
            continue
        src_x1 = x1 - slot.paste_x
        src_y1 = y1 - slot.paste_y
        src_x2 = src_x1 + (x2 - x1)
        src_y2 = src_y1 + (y2 - y1)
        alpha_crop = alpha.crop((src_x1, src_y1, src_x2, src_y2))
        existing = mask.crop((x1, y1, x2, y2))
        merged = ImageChops.lighter(existing, alpha_crop)
        mask.paste(merged, (x1, y1))

    if dilate_px > 0:
        kernel = dilate_px * 2 + 1
        mask = mask.filter(ImageFilter.MaxFilter(kernel))
    return mask


def _count_outside_mask_diff_pixels(candidate: Image.Image, baseline: Image.Image, mask: Image.Image) -> int:
    """
    Count how many pixels differ between candidate and baseline outside mask.
    mask=255 means allowed-to-change area; mask=0 means must stay identical.
    """
    diff = ImageChops.difference(candidate.convert("RGBA"), baseline.convert("RGBA"))
    count = 0
    for diff_px, mask_px in zip(diff.getdata(), mask.getdata()):
        if mask_px == 0 and diff_px != (0, 0, 0, 0):
            count += 1
    return count


def _apply_background_lock(candidate: Image.Image, baseline: Image.Image, mask: Image.Image) -> Image.Image:
    """Keep candidate only inside mask; outside area is restored from baseline."""
    locked = baseline.copy()
    locked.paste(candidate, (0, 0), mask)
    return locked


# ---------------------------------------------------------------------------
# Job management
# ---------------------------------------------------------------------------

def create_job(template_id: str, excel_bytes: bytes) -> dict:
    job_id = uuid.uuid4().hex
    job_dir = OUTPUTS_DIR / job_id
    job_dir.mkdir(parents=True, exist_ok=True)
    excel_path = job_dir / "mapping.xlsx"
    excel_path.write_bytes(excel_bytes)

    with get_conn() as conn:
        conn.execute(
            """
            INSERT INTO jobs (id, template_id, excel_path, status, progress, log_json, output_dir, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                job_id,
                template_id,
                str(excel_path),
                "pending",
                0.0,
                dump_json(["任务已创建"]),
                str(job_dir),
                datetime.utcnow().isoformat(),
            ),
        )
        conn.commit()

    worker = threading.Thread(target=_run_job, args=(job_id,), daemon=True)
    worker.start()
    return {"job_id": job_id, "status": "pending"}


def _load_group_info(template_id: str) -> list[dict]:
    """
    Return all replace groups for the template, with each group's
    "replace" layer records (id, x, y, width, height) resolved from DB.
    """
    with get_conn() as conn:
        group_rows = conn.execute(
            """
            SELECT id, name, region_json, layer_rules_json
            FROM replace_groups
            WHERE template_id = ?
            ORDER BY created_at ASC
            """,
            (template_id,),
        ).fetchall()

        layer_rows = conn.execute(
            "SELECT id, name, x, y, width, height, stack_index FROM layers WHERE template_id = ?",
            (template_id,),
        ).fetchall()

    layer_lookup = {r["id"]: dict(r) for r in layer_rows}

    groups = []
    for row in group_rows:
        rules = json.loads(row["layer_rules_json"] or "[]")
        replace_layers = []
        for rule in rules:
            if rule.get("action", "replace") != "replace":
                continue
            layer_id = rule["layer_id"]
            if layer_id in layer_lookup:
                replace_layers.append(layer_lookup[layer_id])
        replace_layers.sort(key=lambda layer: int(layer.get("stack_index", 0)))
        groups.append(
            {
                "id": row["id"],
                "name": row["name"],
                "replace_layers": replace_layers,
                "all_layer_ids": [r["layer_id"] for r in rules if r.get("action", "replace") == "replace"],
            }
        )
    return groups


def _compute_union_bbox(layers: list[dict]) -> tuple[int, int, int, int]:
    """Return (x, y, w, h) bounding box of all layers."""
    min_x = min(l["x"] for l in layers)
    min_y = min(l["y"] for l in layers)
    max_x2 = max(l["x"] + l["width"] for l in layers)
    max_y2 = max(l["y"] + l["height"] for l in layers)
    return min_x, min_y, max_x2 - min_x, max_y2 - min_y


def _run_job(job_id: str) -> None:
    try:
        with get_conn() as conn:
            job = conn.execute("SELECT * FROM jobs WHERE id = ?", (job_id,)).fetchone()
            if not job:
                return
            tpl = conn.execute(
                "SELECT file_path, preview_path FROM templates WHERE id = ?",
                (job["template_id"],),
            ).fetchone()

        if not tpl:
            _update_job(job_id, "failed", 0.0, ["模板不存在，任务终止"])
            return

        logs: list[str] = ["开始执行批量任务"]
        _update_job(job_id, "running", 0.01, logs)

        job_dir = Path(job["output_dir"])
        outputs_dir = job_dir / "png"
        outputs_dir.mkdir(exist_ok=True)

        # ── 1. Load Excel ──
        wb = load_workbook(job["excel_path"])
        ws = wb.active

        # Extract images from Excel
        extracted_assets_dir = job_dir / "extracted_assets"
        extracted_assets_dir.mkdir(exist_ok=True)
        excel_images: dict[tuple[int, int], Path] = {}

        excel_bytes = Path(job["excel_path"]).read_bytes()
        wps_images_data = _extract_wps_embedded_images(excel_bytes)
        wps_images_paths: dict[str, Path] = {}
        for img_id, img_bytes in wps_images_data.items():
            path = extracted_assets_dir / f"wps_{img_id}.png"
            path.write_bytes(img_bytes)
            wps_images_paths[img_id] = path

        if hasattr(ws, "_images"):
            for i_idx, img in enumerate(ws._images):
                try:
                    # row, col are 1-indexed here for our convenience
                    r = img.anchor._from.row + 1
                    c = img.anchor._from.col + 1
                    img_data = img._data()
                    # Determine extension (usually png or jpeg)
                    ext = ".png"
                    if hasattr(img, "format") and img.format:
                        ext = f".{img.format.lower()}"
                    elif hasattr(img, "path") and img.path:
                        ext = Path(img.path).suffix

                    filename = f"excel_{r}_{c}_{i_idx}{ext}"
                    path = extracted_assets_dir / filename
                    path.write_bytes(img_data)
                    excel_images[(r, c)] = path
                except Exception as e:
                    logs.append(f"提取 Excel 图片失败 (index {i_idx}): {e}")

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            _update_job(job_id, "failed", 0.0, logs + ["Excel 为空"])
            return
        headers = [str(x).strip() if x else "" for x in rows[0]]
        group_col_map = _parse_group_columns(headers)
        logs.append(f"Excel 表头: {headers}")
        logs.append(f"解析到替换组列: {group_col_map}")

        # ── 2. Load replace group definitions ──
        template_id = job["template_id"]
        groups = _load_group_info(template_id)
        groups_by_name = {g["name"]: g for g in groups}
        logs.append(f"DB 替换组: {list(groups_by_name.keys())}")
        for g in groups:
            layer_desc = [(l["name"], l["x"], l["y"], l["width"], l["height"]) for l in g["replace_layers"]]
            logs.append(f"  组 '{g['name']}': {len(g['replace_layers'])} 个图层 {layer_desc}")

        # Warn if no matching groups
        matched = [gn for gn in group_col_map if gn in groups_by_name]
        unmatched_excel = [gn for gn in group_col_map if gn not in groups_by_name]
        unmatched_db = [gn for gn in groups_by_name if gn not in group_col_map]
        if unmatched_excel:
            logs.append(f"警告: Excel 中的替换组在 DB 中不存在: {unmatched_excel}")
        if unmatched_db:
            logs.append(f"警告: DB 中的替换组在 Excel 中无对应列: {unmatched_db}")
        logs.append(f"匹配的替换组: {matched}")

        if not matched:
            _update_job(job_id, "failed", 0.0, logs + ["无匹配的替换组，任务终止"])
            return

        # ── 3. Collect all layer IDs to hide ──
        all_replace_ids: list[str] = []
        for g in groups:
            all_replace_ids.extend(g["all_layer_ids"])
        all_replace_ids = list(dict.fromkeys(all_replace_ids))
        replace_layer_id_set = set(all_replace_ids)
        logs.append(f"需隐藏的图层数: {len(all_replace_ids)}")

        # ── 4. Open PSD and build base canvas ONCE ──
        psd_path = tpl["file_path"]
        base_canvas: Image.Image | None = None
        ordered_leaf_layers: list = []
        layer_id_by_object_id: dict[int, str] = {}
        ordered_replace_layer_ids: list[str] = []
        ordered_replace_layer_id_set: set[str] = set()
        precise_render_enabled = False
        try:
            psd_obj = PSDImage.open(psd_path)
            layer_map = build_layer_map(psd_obj, template_id)
            logs.append(f"PSD loaded, layer_map matched {len(layer_map)} layers")

            mapped_ids = [lid for lid in all_replace_ids if lid in layer_map]
            unmapped_ids = [lid for lid in all_replace_ids if lid not in layer_map]
            logs.append(f"Replace layers matched in PSD: {len(mapped_ids)}/{len(all_replace_ids)}")
            if unmapped_ids:
                logs.append(f"Warning: unmatched PSD layer IDs: {unmapped_ids}")

            base_canvas, composite_meta = composite_without_layers(psd_obj, all_replace_ids, layer_map)
            logs.append(
                f"Base canvas generated: size={base_canvas.size}, "
                f"composite_mode={composite_meta['composite_mode']}, "
                f"requested_hide_count={composite_meta['requested_hide_count']}, "
                f"effective_hide_count={composite_meta['effective_hide_count']}, "
                f"pixel_changed={composite_meta['pixel_changed']}"
            )
            if composite_meta.get("fallback_used"):
                raise RuntimeError(
                    "force composite fallback requested: "
                    f"{composite_meta.get('composite_error') or 'unknown error'}"
                )

            _flatten_leaf_layers(psd_obj, ordered_leaf_layers)
            layer_id_by_object_id = {
                id(psd_layer): layer_id
                for layer_id, psd_layer in layer_map.items()
            }
            ordered_replace_layer_ids = [
                layer_id_by_object_id.get(id(layer))
                for layer in ordered_leaf_layers
                if layer_id_by_object_id.get(id(layer)) in replace_layer_id_set
            ]
            ordered_replace_layer_id_set = set(ordered_replace_layer_ids)
            if ordered_leaf_layers:
                precise_render_enabled = True
                logs.append(
                    f"layer-aware render prepared: leaf_layers={len(ordered_leaf_layers)}, "
                    f"ordered_replace_layers={len(ordered_replace_layer_ids)}"
                )
            else:
                logs.append("layer-aware render disabled: no leaf layers found")
        except Exception as exc:
            logs.append(f"PSD composite failed, using preview fallback: {exc}")

        if base_canvas is None:
            preview_path = tpl["preview_path"]
            if Path(preview_path).exists():
                base_canvas = Image.open(preview_path).convert("RGBA")
                logs.append(f"使用预览图作为底图: {base_canvas.size}")
            else:
                _update_job(job_id, "failed", 0.0, logs + ["无法生成基础画布"])
                return

        _update_job(job_id, "running", 0.05, logs)

        # ── 5. Process each Excel row ──
        body_rows = rows[1:]
        total = max(1, len(body_rows))
        generated_count = 0
        precise_render_fallback_logged = False

        for index, row_data_tuple in enumerate(body_rows):
            if row_data_tuple is None:
                continue
            r_idx = index + 2 # 1-based, 1 is header
            data = {
                headers[i]: ("" if row_data_tuple[i] is None else str(row_data_tuple[i]).strip())
                for i in range(min(len(headers), len(row_data_tuple)))
            }
            output_name = data.get("output_name") or f"output_{index + 1:03d}"

            row_log_parts: list[str] = []
            row_layer_replacements: dict[str, SlotResult] = {}
            top_overlay_slots: list[SlotResult] = []

            dispimg_re = re.compile(r'DISPIMG\("(.+?)"')

            for group_name, col_list in group_col_map.items():
                group_def = groups_by_name.get(group_name)
                if not group_def:
                    row_log_parts.append(f"组'{group_name}'无匹配")
                    continue

                # Collect assets (either from filename or excel image)
                cutout_imgs: list[Image.Image] = []
                for c_idx_in_headers, col in enumerate(headers, start=1):
                    if col not in col_list:
                        continue

                    # 即使 asset_name 有值，如果是 __IMAGE_AT_... 占位符，也应视为需要从 Excel 读取
                    asset_name = data.get(col, "").strip()
                    is_placeholder = asset_name.startswith("__IMAGE_AT_") and asset_name.endswith("__")
                    is_wps_placeholder = asset_name.startswith("__WPS_IMAGE_") and asset_name.endswith("__")

                    # Check if we have an image at this (r_idx, c_idx)
                    if (not asset_name or is_placeholder) and (r_idx, c_idx_in_headers) in excel_images:
                        img_path = excel_images[(r_idx, c_idx_in_headers)]
                        try:
                            img = Image.open(img_path).convert("RGBA")
                            cutout_imgs.append(img)
                        except Exception as exc:
                            row_log_parts.append(f"Excel 图片'{img_path.name}'解析失败:{exc}")
                    elif is_wps_placeholder:
                        img_id = asset_name[len("__WPS_IMAGE_"):-2]
                        if img_id in wps_images_paths:
                            img_path = wps_images_paths[img_id]
                            try:
                                img = Image.open(img_path).convert("RGBA")
                                cutout_imgs.append(img)
                            except Exception as exc:
                                row_log_parts.append(f"WPS 图片'{img_id}'解析失败:{exc}")
                    elif asset_name and not is_placeholder and not is_wps_placeholder:
                        # Fallback for manual DISPIMG string if not already converted by import_excel_and_validate
                        m = dispimg_re.search(asset_name)
                        if m:
                            img_id = m.group(1)
                            if img_id in wps_images_paths:
                                img_path = wps_images_paths[img_id]
                                try:
                                    img = Image.open(img_path).convert("RGBA")
                                    cutout_imgs.append(img)
                                    continue # Skip the filename fallback
                                except Exception:
                                    pass

                        try:
                            # Original logic for filenames
                            cutout_path = cutout_asset(asset_name)
                            img = Image.open(cutout_path).convert("RGBA")
                            cutout_imgs.append(img)
                        except Exception as exc:
                            row_log_parts.append(f"素材'{asset_name}'失败:{exc}")

                if not cutout_imgs:
                    continue

                replace_layers = group_def["replace_layers"]
                n_imgs = len(cutout_imgs)
                n_layers = len(replace_layers)

                if n_layers > 0 and n_imgs == n_layers:
                    detail_parts: list[str] = []
                    for img, layer_info in zip(cutout_imgs, replace_layers):
                        orig_size = img.size
                        slot = _fit_image_to_slot(
                            img,
                            layer_info["x"],
                            layer_info["y"],
                            layer_info["width"],
                            layer_info["height"],
                        )
                        row_layer_replacements[layer_info["id"]] = slot
                        detail_parts.append(
                            f"  [{layer_info['name']} slot=({layer_info['x']},{layer_info['y']},"
                            f"{layer_info['width']}x{layer_info['height']}) "
                            f"src={orig_size} -> trimmed+scaled={slot.img.size} "
                            f"paste=({slot.paste_x},{slot.paste_y})]"
                        )
                    row_log_parts.append(
                        f"组'{group_name}': 1:1替换 {n_imgs}张\n" + "\n".join(detail_parts)
                    )
                else:
                    if replace_layers:
                        bx, by, bw, bh = _compute_union_bbox(replace_layers)
                    else:
                        bx, by = 0, 0
                        bw, bh = base_canvas.width, base_canvas.height

                    slot_results = _grid_layout(cutout_imgs, bx, by, bw, bh)
                    detail_parts = []
                    for idx_s, slot in enumerate(slot_results):
                        if idx_s < n_layers:
                            row_layer_replacements[replace_layers[idx_s]["id"]] = slot
                        else:
                            top_overlay_slots.append(slot)
                        detail_parts.append(
                            f"  [slot{idx_s} scaled={slot.img.size} paste=({slot.paste_x},{slot.paste_y})]"
                        )
                    row_log_parts.append(
                        f"组'{group_name}': 网格布局 {n_imgs}张 bbox=({bx},{by},{bw}x{bh})\n"
                        + "\n".join(detail_parts)
                    )

            unmatched_slots_for_top: list[SlotResult] = []
            for layer_id, slot in row_layer_replacements.items():
                if layer_id not in ordered_replace_layer_id_set:
                    unmatched_slots_for_top.append(slot)
            if unmatched_slots_for_top:
                top_overlay_slots.extend(unmatched_slots_for_top)
                row_log_parts.append(f"未映射替换层回退顶层渲染: {len(unmatched_slots_for_top)}")

            if precise_render_enabled:
                try:
                    layer_canvas = _render_layer_aware_canvas(
                        ordered_layers=ordered_leaf_layers,
                        layer_id_by_object_id=layer_id_by_object_id,
                        replace_layer_ids=replace_layer_id_set,
                        replacement_slots=row_layer_replacements,
                        canvas_size=base_canvas.size,
                    )
                except Exception as exc:
                    precise_render_enabled = False
                    if not precise_render_fallback_logged:
                        logs.append(f"layer-aware render failed, fallback to top-overlay mode: {exc}")
                        precise_render_fallback_logged = True
                    layer_canvas = base_canvas.copy()
                    for slot in row_layer_replacements.values():
                        layer_canvas.alpha_composite(slot.img, (slot.paste_x, slot.paste_y))
            else:
                layer_canvas = base_canvas.copy()
                for slot in row_layer_replacements.values():
                    layer_canvas.alpha_composite(slot.img, (slot.paste_x, slot.paste_y))

            for slot in top_overlay_slots:
                layer_canvas.alpha_composite(slot.img, (slot.paste_x, slot.paste_y))

            editable_slots = list(row_layer_replacements.values()) + top_overlay_slots
            if editable_slots:
                edit_mask = _build_slots_mask(base_canvas.size, editable_slots)
                outside_before = _count_outside_mask_diff_pixels(layer_canvas, base_canvas, edit_mask)
                canvas = _apply_background_lock(layer_canvas, base_canvas, edit_mask)
                outside_after = _count_outside_mask_diff_pixels(canvas, base_canvas, edit_mask)
                row_log_parts.append(
                    f"背景锁定: outside_diff_before={outside_before}, outside_diff_after={outside_after}"
                )
            else:
                canvas = base_canvas.copy()

            out_file = outputs_dir / f"{output_name}.png"
            canvas.save(out_file)
            generated_count += 1
            logs.append(f"[{index+1}/{total}] {output_name}: {'; '.join(row_log_parts) or '无操作'}")
            progress = 0.05 + 0.95 * (index + 1) / total
            _update_job(job_id, "running", progress, logs)

        logs.append(f"任务完成，共生成 {generated_count} 张图片")
        _update_job(job_id, "completed", 1.0, logs)

    except Exception as exc:
        import traceback
        tb = traceback.format_exc()
        _update_job(job_id, "failed", 0.0, [f"任务异常终止: {exc}", tb])


def _update_job(job_id: str, status: str, progress: float, logs: list[str]) -> None:
    with get_conn() as conn:
        conn.execute(
            "UPDATE jobs SET status = ?, progress = ?, log_json = ? WHERE id = ?",
            (status, progress, json.dumps(logs, ensure_ascii=False), job_id),
        )
        conn.commit()


def get_job(job_id: str) -> dict:
    with get_conn() as conn:
        row = conn.execute("SELECT * FROM jobs WHERE id = ?", (job_id,)).fetchone()
    if not row:
        raise ValueError("任务不存在")
    return {
        "id": row["id"],
        "status": row["status"],
        "progress": row["progress"],
        "logs": json.loads(row["log_json"] or "[]"),
        "output_dir": row["output_dir"],
    }


def list_output_images(job_id: str) -> list[str]:
    job = get_job(job_id)
    png_dir = Path(job["output_dir"]) / "png"
    if not png_dir.exists():
        return []
    return sorted(f.name for f in png_dir.glob("*.png"))


def build_output_zip(job_id: str) -> str:
    job = get_job(job_id)
    output_dir = Path(job["output_dir"]) / "png"
    zip_base = Path(job["output_dir"]) / "outputs"
    zip_path = shutil.make_archive(str(zip_base), "zip", root_dir=output_dir)
    return zip_path


def export_psd_basic(job_id: str) -> dict:
    job = get_job(job_id)
    preview_psd = Path(job["output_dir"]) / "README_PSD_EXPORT.txt"
    preview_psd.write_text(
        "当前版本已完成 PNG 主链路与批处理。\n"
        "PSD 高保真导出正在增强阶段，建议先使用 PNG 产物。\n",
        encoding="utf-8",
    )
    return {
        "status": "partial",
        "message": "已完成 PSD 导出增强接口与占位产物，后续可扩展为分层 PSD 真写回。",
        "artifact": str(preview_psd),
    }
