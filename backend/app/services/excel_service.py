from __future__ import annotations

import io
import json
import re
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

from openpyxl import Workbook, load_workbook

from ..config import ASSETS_DIR
from ..db import get_conn


def export_excel_template(template_id: str) -> bytes:
    """
    Generate an Excel template where each replacement group gets N columns
    (group_{name}_1 ... group_{name}_N), where N is the number of "replace"
    layer rules in that group.
    """
    with get_conn() as conn:
        groups = conn.execute(
            """
            SELECT name, layer_rules_json
            FROM replace_groups
            WHERE template_id = ?
            ORDER BY created_at ASC
            """,
            (template_id,),
        ).fetchall()
        tpl = conn.execute(
            "SELECT name FROM templates WHERE id = ?",
            (template_id,),
        ).fetchone()

    if not tpl:
        raise ValueError("模板不存在")
    if not groups:
        raise ValueError("请先创建替换组")

    wb = Workbook()
    ws = wb.active
    ws.title = "mapping"

    group_headers: list[str] = []
    for row in groups:
        layer_rules = json.loads(row["layer_rules_json"] or "[]")
        replace_count = sum(1 for r in layer_rules if r.get("action", "replace") == "replace")
        count = max(1, replace_count)
        for i in range(1, count + 1):
            group_headers.append(f"group_{row['name']}_{i}")

    headers = ["output_name", "template_name"] + group_headers + ["notes"]
    ws.append(headers)

    example_row = ["example_001", tpl["name"]] + ["" for _ in group_headers] + ["填写商品文件名，不带后缀"]
    ws.append(example_row)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _parse_group_columns(headers: list[str]) -> dict[str, list[str]]:
    """
    Parse multi-column group headers like group_{name}_1, group_{name}_2, ...
    Also accepts legacy single-column format group_{name} (without numeric suffix).
    Returns dict: group_name -> [col_header, ...]  ordered by suffix number.
    """
    pattern = re.compile(r"^group_(.+?)_(\d+)$")
    legacy_pattern = re.compile(r"^group_(.+)$")

    groups: dict[str, list[tuple[int, str]]] = {}

    for h in headers:
        m = pattern.match(h)
        if m:
            gname = m.group(1)
            idx = int(m.group(2))
            groups.setdefault(gname, []).append((idx, h))
            continue
        m2 = legacy_pattern.match(h)
        if m2:
            gname = m2.group(1)
            # treat as index 1
            groups.setdefault(gname, []).append((1, h))

    return {
        gname: [col for _, col in sorted(cols)]
        for gname, cols in groups.items()
    }


def _extract_wps_embedded_images(file_bytes: bytes) -> dict[str, bytes]:
    """
    Extracts WPS/Excel 'cell images' (DISPIMG) that openpyxl misses.
    Returns a dict mapping ID (e.g. ID_...) to image bytes.
    """
    images = {}
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes), 'r') as z:
            if 'xl/cellimages.xml' not in z.namelist():
                return images

            cell_images_xml = z.read('xl/cellimages.xml')
            root = ET.fromstring(cell_images_xml)
            ns = {
                'etc': 'http://www.wps.cn/officeDocument/2017/etCustomData',
                'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
                'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
                'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
            }

            for pic in root.findall('.//xdr:pic', ns):
                name_elem = pic.find('.//xdr:cNvPr', ns)
                blip_elem = pic.find('.//a:blip', ns)
                if name_elem is not None and blip_elem is not None:
                    img_id = name_elem.get('name')
                    r_id = blip_elem.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
                    if img_id and r_id:
                        rels_path = 'xl/_rels/cellimages.xml.rels'
                        if rels_path in z.namelist():
                            rels_xml = z.read(rels_path)
                            rels_root = ET.fromstring(rels_xml)
                            for rel in rels_root.findall('{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                                if rel.get('Id') == r_id:
                                    target = rel.get('Target')
                                    img_path = f'xl/{target}' if not target.startswith('xl/') else target
                                    if img_path in z.namelist():
                                        images[img_id] = z.read(img_path)
    except Exception:
        pass
    return images


def import_excel_and_validate(template_id: str, file_bytes: bytes) -> dict:
    wb = load_workbook(io.BytesIO(file_bytes))
    ws = wb.active

    # 提取图片
    images_map = {} # (row, col) -> image_data
    if hasattr(ws, "_images"):
        for img in ws._images:
            # openpyxl 3.0+ images have an anchor._from property
            # row and col are 0-indexed in anchor
            row = img.anchor._from.row + 1
            col = img.anchor._from.col + 1
            images_map[(row, col)] = img

    # 提取 WPS 嵌入式图片 (DISPIMG)
    wps_images = _extract_wps_embedded_images(file_bytes)
    dispimg_re = re.compile(r'DISPIMG\("(.+?)"')

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel 为空")
    headers = [str(x).strip() if x is not None else "" for x in rows[0]]

    group_col_map = _parse_group_columns(headers)
    group_names = list(group_col_map.keys())

    with get_conn() as conn:
        existing_groups = conn.execute(
            "SELECT name FROM replace_groups WHERE template_id = ?",
            (template_id,),
        ).fetchall()
    valid_group_set = {x["name"] for x in existing_groups}
    unknown_groups = [g for g in group_names if g not in valid_group_set]

    # asset_names = {Path(p.name).stem for p in ASSETS_DIR.glob("*.*")}
    parsed_rows: list[dict] = []
    missing_assets: set[str] = set()

    for r_idx, row in enumerate(rows[1:], start=2):
        if row is None:
            continue
        item: dict[str, str] = {}
        for c_idx, key in enumerate(headers, start=1):
            val = row[c_idx-1] if c_idx-1 < len(row) else None
            item[key] = "" if val is None else str(val).strip()

            # 如果是组列且值为空，检查是否有图片
            is_group_col = any(key in cols for cols in group_col_map.values())
            if is_group_col:
                if not item[key]:
                    if (r_idx, c_idx) in images_map:
                        item[key] = f"__IMAGE_AT_{r_idx}_{c_idx}__"
                elif 'DISPIMG' in item[key]:
                    m = dispimg_re.search(item[key])
                    if m:
                        img_id = m.group(1)
                        if img_id in wps_images:
                            item[key] = f"__WPS_IMAGE_{img_id}__"

        if any(item.values()):
            parsed_rows.append(item)
            for col_list in group_col_map.values():
                for col in col_list:
                    val = item.get(col, "").strip()
                    # 现在不强制要求在 ASSETS_DIR 中，但如果不是特殊图片占位符且不为空，则可能是旧逻辑或填写错误
                    # 如果没有占位符也没有文件名，标记为缺失（可选，取决于是否允许留空）
                    if not val:
                        # 暂时不标记为缺失，允许跳过某些层
                        pass

    return {
        "rows": parsed_rows,
        "missing_assets": sorted(missing_assets),
        "unknown_groups": unknown_groups,
        "group_col_map": {k: v for k, v in group_col_map.items()},
    }
